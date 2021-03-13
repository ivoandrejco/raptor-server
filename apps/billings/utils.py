import os, sys
import django
from django.conf import settings
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0,os.path.join(BASE_DIR))

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'raptor.settings')
django.setup()

from django.db.utils import IntegrityError

from openpyxl import load_workbook

from billings.models import Claim
from doctors.models import ProviderNumber

amounts={110:134.3,116:67.2}
def import_claims_from_excel_old(filename,field_ranges):
    wb          = load_workbook(filename)
    # filename should be provider number
    basename    = os.path.basename(filename)
    pnumber     = basename.split(".")[0]
    billed_by   = ProviderNumber.objects.get(number=pnumber)
    for sheet_name in field_ranges:
        print(f'Sheet name: {sheet_name} {field_ranges[sheet_name]}')
        sheet   = wb[sheet_name]
        ranges  = field_ranges[sheet_name]
        for field_range in field_ranges[sheet_name]:
            dt = None
            for c, d in sheet[field_range[0]:field_range[1]]:
                if not c.value:
                    dt = d.value
                    continue
                name = d.value.split(" ")
                amount = amounts[int(c.value)]
                try:
                    claim   = Claim(fname=name[0],lname=name[1],billed_on=dt,billed_by=billed_by,code=c.value,amount=amount)
                    claim.save()
                except IntegrityError:
                    print(f'Integrity error {name[0]} {name[1]}')

                print(f'{name[0]} {name[1]} billed_on={dt} billed_by={billed_by} code={c.value} amount={amount}')

def import_claims_from_excel(filename):
    wb          = load_workbook(filename)
    i = 1
    for sheet in wb:
        print(f'Reading sheet: {sheet.title}')
        billed_by   = ProviderNumber.objects.get(number=sheet.title)
        j = 1
        for a,b,c in sheet['A1':'C200']:
            if not b.value:
                continue
            name    = c.value.split(" ",1)
            code    = int(b.value)
            amount  = amounts[code]
            print(f'({i} - {j}) - Values: {a.value} {name[0]} {name[1]} {code} {amount} {billed_by}')
            i += 1
            j += 1
#                try:
#                    claim   = Claim(fname=name[0],lname=name[1],billed_on=dt,billed_by=billed_by,code=c.value,amount=amount)
#                    claim.save()
#                except IntegrityError:
#                    print(f'Integrity error {name[0]} {name[1]}')

def import_payments_from_excel(filename):
    wb          = load_workbook(filename)
    i = 1
    for sheet in wb:
        print(f'Reading sheet: {sheet.title}')
        billed_by   = ProviderNumber.objects.get(number=sheet.title)
        j = 1
        for a,b,c,d,e in sheet['A1':'E205']:
            if not b.value:
                continue
            name    = b.value.split(" ",1)
            code    = int(c.value)
            amount  = float(d.value)
            print(f'({i} - {j}) - Values: {a.value} {name[0]} {name[1]} {code} {amount} {e.value} {billed_by}')
            i += 1
            j += 1
#                try:
#                    claim   = Claim(fname=name[0],lname=name[1],billed_on=dt,billed_by=billed_by,code=c.value,amount=amount)
#                    claim.save()
#                except IntegrityError:
#                    print(f'Integrity error {name[0]} {name[1]}')

#                print(f'{name[0]} {name[1]} billed_on={dt} billed_by={billed_by} code={c.value} amount={amount}')


if __name__ == '__main__':
    import_claims_from_excel('claims_1020.xlsx')
#    import_payments_from_excel('payments.xlsx')
